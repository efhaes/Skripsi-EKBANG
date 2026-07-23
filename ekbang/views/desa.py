from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from ekbang.proses.decorators import role_required
from django.db.models import Q
from ekbang.models import (
    Desa, Warga, HasilSAW, PengajuanBLT, NormalisasiSAW, KuotaKPM,
    Kriteria, penilaianwarga, NormalisasiSAWDetail
)
from ekbang.proses.saw import hitung_saw
from ekbang.forms import WargaForm, PengajuanBLTForm, KuotaKPMForm
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

@login_required
@role_required('desa')
def dashboard_desa(request):
    desa = request.user.desa

    context = {
        'total_warga': Warga.objects.filter(desa=desa, is_deleted=False).count(),
        'total_hasil': HasilSAW.objects.filter(
            normalisasi__desa=desa,
            normalisasi__is_active=True
        ).count(),
        'pengajuan': PengajuanBLT.objects.filter(desa=desa).order_by('-created_at')[:5],
        'total_pengajuan': PengajuanBLT.objects.filter(desa=desa).count(),
        'pengajuan_pending': PengajuanBLT.objects.filter(
            desa=desa,
            status='pending'
        ).count(),
    }
    return render(request, 'desa/dashboard.html', context)


@login_required
@role_required('desa')
def warga_list(request):
    desa = request.user.desa
    query = request.GET.get('q', '')

    warga = Warga.objects.filter(
        desa=desa,
        is_deleted=False
    )

    if query:
        warga = warga.filter(
            Q(nama__icontains=query) |
            Q(nik__icontains=query)
        )

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string(
            'desa/partials/warga_rows.html',
            {'warga': warga}
        )

        cards_html = render_to_string(
            'desa/partials/warga_cards.html',
            {'warga': warga}
        )

        return JsonResponse({
            'html': html,
            'cards_html': cards_html
        })

    return render(request, 'desa/warga_list.html', {
        'warga': warga,
        'query': query
    })


@login_required
@role_required('desa')
def warga_tambah(request):
    desa = request.user.desa

    form = WargaForm(
        request.POST or None,
        desa=desa
    )

    if request.method == "POST" and form.is_valid():
        warga = form.save(commit=False)
        warga.desa = desa
        warga.save()

        form.save_nilai_kriteria(warga)

        messages.success(
            request,
            "Data warga berhasil ditambahkan."
        )

        return redirect("desa_warga_list")

    return render(
        request,
        "desa/warga_form.html",
        {
            "form": form,
            "mode": "tambah",
        },
    )


@login_required
@role_required('desa')
def warga_edit(request, id):
    desa = request.user.desa

    warga = get_object_or_404(
        Warga,
        id=id,
        desa=desa,
        is_deleted=False
    )

    form = WargaForm(
        request.POST or None,
        instance=warga,
        desa=desa
    )

    if request.method == "POST" and form.is_valid():
        warga = form.save()

        form.save_nilai_kriteria(warga)

        messages.success(
            request,
            "Data warga berhasil diupdate."
        )

        return redirect("desa_warga_list")

    return render(
        request,
        "desa/warga_form.html",
        {
            "form": form,
            "mode": "edit",
            "warga": warga,
        },
    )


@login_required
@role_required('desa')
def warga_hapus(request, id):
    desa = request.user.desa

    warga = get_object_or_404(
        Warga,
        id=id,
        desa=desa,
        is_deleted=False
    )

    warga.is_deleted = True
    warga.deleted_at = timezone.now()
    warga.save(
        update_fields=[
            "is_deleted",
            "deleted_at",
        ]
    )

    messages.success(
        request,
        "Data warga berhasil dihapus."
    )

    return redirect("desa_warga_list")


@login_required
@role_required('desa')
def warga_detail(request, id):
    desa = request.user.desa

    warga = get_object_or_404(
        Warga,
        id=id,
        desa=desa,
        is_deleted=False,
    )

    hasil_saw = (
        HasilSAW.objects
        .select_related(
            "normalisasi",
            "normalisasi__desa",
            "normalisasi__warga",
        )
        .filter(
            normalisasi__warga=warga,
            normalisasi__desa=desa,
            normalisasi__is_active=True,
        )
        .first()
    )

    if hasil_saw:
        normalisasi = hasil_saw.normalisasi
    else:
        normalisasi = (
            NormalisasiSAW.objects
            .filter(
                warga=warga,
                desa=desa,
                is_active=True,
            )
            .first()
        )

    detail_normalisasi = (
        NormalisasiSAWDetail.objects
        .filter(normalisasi=normalisasi)
        .select_related("kriteria")
        if normalisasi
        else []
    )

    nilai_kriteria = (
        penilaianwarga.objects
        .filter(warga=warga)
        .select_related(
            "kriteria",
            "subkriteria",
        )
    )

    context = {
        "warga": warga,
        "hasil_saw": hasil_saw,
        "normalisasi": normalisasi,
        "detail_normalisasi": detail_normalisasi,
        "nilai_kriteria": nilai_kriteria,
    }

    return render(
        request,
        "desa/warga_detail.html",
        context,
    )


def _parse_tahun(tahun_str):
    tahun_str = (tahun_str or "").strip()

    if not tahun_str:
        return None

    try:
        tahun = int(tahun_str)
    except ValueError:
        return None

    if tahun < 2000 or tahun > 2100:
        return None

    return tahun


from ekbang.models import Kriteria, Subkriteria
from ekbang.forms import KriteriaForm, SubkriteriaForm


BATAS_TOTAL_BOBOT = 1.0
TOLERANSI = 0.001  # antisipasi floating point, mis. 0.9999999999


@login_required
@role_required('desa')
def kriteria_list(request):
    desa = request.user.desa
    kriteria = Kriteria.objects.filter(desa=desa).order_by('urutan', 'id')

    total_bobot = round(sum(k.bobot for k in kriteria if k.aktif), 2)

    return render(request, 'desa/kriteria_list.html', {
        'kriteria': kriteria,
        'total_bobot': total_bobot,
        'bobot_maksimum': total_bobot >= BATAS_TOTAL_BOBOT - TOLERANSI,
    })

@login_required
@role_required('desa')
def kriteria_tambah(request):
    desa = request.user.desa

    total_bobot_aktif = sum(
        k.bobot for k in Kriteria.objects.filter(desa=desa, aktif=True)
    )

    if total_bobot_aktif >= 1:
        messages.warning(
            request,
            'Total bobot kriteria sudah Mencapai 1.00. Harap Hapus atau kurangi bobot kriteria lain dulu.'
        )
        return redirect('desa_kriteria_list')

    form = KriteriaForm(request.POST or None, desa=desa)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Kriteria berhasil ditambahkan')
        return redirect('desa_kriteria_list')

    return render(request, 'desa/kriteria_form.html', {'form': form, 'mode': 'tambah'})


@login_required
@role_required('desa')
def kriteria_edit(request, id):
    desa = request.user.desa
    kriteria = get_object_or_404(Kriteria, id=id, desa=desa)
    form = KriteriaForm(request.POST or None, instance=kriteria, desa=desa)

    if request.method == 'POST' and form.is_valid():
        # Total bobot aktif TANPA kriteria ini (karena mau diganti nilainya)
        total_bobot_lain = round(
            sum(
                k.bobot for k in Kriteria.objects.filter(desa=desa, aktif=True).exclude(id=kriteria.id)
            ), 2
        )
        bobot_baru = form.cleaned_data['bobot']
        aktif_baru = form.cleaned_data.get('aktif', kriteria.aktif)

        if aktif_baru and round(total_bobot_lain + bobot_baru, 2) > BATAS_TOTAL_BOBOT + TOLERANSI:
            messages.error(
                request,
                f'Bobot {bobot_baru:.2f} membuat total bobot melebihi 1.00 '
                f'(kriteria lain totalnya {total_bobot_lain:.2f}). Harap kurangi bobot kriteria lain terlebih dahulu.'
            )
            return render(request, 'desa/kriteria_form.html', {
                'form': form, 'mode': 'edit', 'kriteria': kriteria
            })

        form.save()
        messages.success(request, 'Kriteria berhasil diupdate')
        return redirect('desa_kriteria_list')

    return render(request, 'desa/kriteria_form.html', {
        'form': form, 'mode': 'edit', 'kriteria': kriteria
    })


@login_required
@role_required('desa')
def kriteria_hapus(request, id):
    desa = request.user.desa
    kriteria = get_object_or_404(Kriteria, id=id, desa=desa)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Method tidak diizinkan.'}, status=405)
        return redirect('desa_kriteria_list')

    if kriteria.nilai_warga.exists():
        msg = (
            'Kriteria ini sudah dipakai untuk penilaian warga dan tidak bisa dihapus. '
            'Nonaktifkan saja kriterianya.'
        )
        if is_ajax:
            return JsonResponse({'success': False, 'message': msg}, status=400)
        messages.warning(request, msg)
        return redirect('desa_kriteria_list')

    nama = kriteria.nama
    kriteria.delete()

    if is_ajax:
        total_bobot = round(
            sum(k.bobot for k in Kriteria.objects.filter(desa=desa, aktif=True)), 2
        )
        return JsonResponse({
            'success': True,
            'message': f'Kriteria "{nama}" berhasil dihapus.',
            'total_bobot': total_bobot,
            'bobot_maksimum': total_bobot >= BATAS_TOTAL_BOBOT - TOLERANSI,
        })

    messages.success(request, 'Kriteria berhasil dihapus')
    return redirect('desa_kriteria_list')   

@login_required
@role_required('desa')
def kriteria_toggle_aktif(request, id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method tidak diizinkan.'}, status=405)

    desa = request.user.desa
    kriteria = get_object_or_404(Kriteria, id=id, desa=desa)

    aktif_baru = not kriteria.aktif

    # Kalau mau diaktifkan, cek dulu batas total bobot
    if aktif_baru:
        total_bobot_lain = round(
            sum(
                k.bobot for k in Kriteria.objects.filter(desa=desa, aktif=True).exclude(id=kriteria.id)
            ), 2
        )
        if round(total_bobot_lain + kriteria.bobot, 2) > BATAS_TOTAL_BOBOT + TOLERANSI:
            return JsonResponse({
                'success': False,
                'message': (
                    f'Tidak bisa mengaktifkan "{kriteria.nama}". '
                    f'Total bobot kriteria aktif lain sudah {total_bobot_lain:.2f}, '
                    f'ditambah bobot kriteria ini ({kriteria.bobot:.2f}) akan melebihi 1.00.'
                )
            }, status=400)

    kriteria.aktif = aktif_baru
    kriteria.save(update_fields=['aktif'])

    total_bobot = round(
        sum(k.bobot for k in Kriteria.objects.filter(desa=desa, aktif=True)), 2
    )

    return JsonResponse({
        'success': True,
        'aktif': kriteria.aktif,
        'total_bobot': total_bobot,
        'bobot_maksimum': total_bobot >= BATAS_TOTAL_BOBOT - TOLERANSI,
    })
# =========================================================
# SUBKRITERIA (nested di dalam satu kriteria)
# =========================================================

@login_required
@role_required('desa')
def subkriteria_list(request):
    desa = request.user.desa

    kriteria_filter = request.GET.get('kriteria', '').strip()

    kriteria_list = (
        Kriteria.objects
        .filter(desa=desa)
        .prefetch_related('subkriteria')
        .order_by('urutan', 'nama')
    )

    if kriteria_filter:
        kriteria_list = kriteria_list.filter(id=kriteria_filter)

    return render(request, 'desa/subkriteria_list.html', {
        'kriteria_list': kriteria_list,
        'kriteria_filter': kriteria_filter,
    })

@login_required
@role_required('desa')
def subkriteria_tambah(request):
    desa = request.user.desa
    form = SubkriteriaForm(request.POST or None, desa=desa)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Subkriteria berhasil ditambahkan')
        return redirect('desa_subkriteria_list')

    return render(request, 'desa/subkriteria_form.html', {'form': form, 'mode': 'tambah'})


@login_required
@role_required('desa')
def subkriteria_edit(request, id):
    desa = request.user.desa
    subkriteria = get_object_or_404(Subkriteria, id=id, kriteria__desa=desa)
    form = SubkriteriaForm(request.POST or None, instance=subkriteria, desa=desa)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Subkriteria berhasil diupdate')
        return redirect('desa_subkriteria_list')

    return render(request, 'desa/subkriteria_form.html', {
        'form': form, 'mode': 'edit', 'subkriteria': subkriteria
    })


@login_required
@role_required('desa')
def subkriteria_hapus(request, id):
    desa = request.user.desa
    subkriteria = get_object_or_404(Subkriteria, id=id, kriteria__desa=desa)

    if request.method == 'POST':
        if subkriteria.nilai_warga.exists():
            messages.warning(
                request,
                'Subkriteria ini sudah dipakai untuk penilaian warga dan tidak bisa dihapus.'
            )
            return redirect('desa_subkriteria_list')

        subkriteria.delete()
        messages.success(request, 'Subkriteria berhasil dihapus')

    return redirect('desa_subkriteria_list')


@login_required
@role_required('desa')
def proses_saw_view(request):
    desa = request.user.desa
    if request.method != 'POST':
        return redirect('desa_proses_saw')
    jumlah_kpm = request.POST.get('jumlah_kpm', '').strip() 
    tahap = request.POST.get('tahap', '').strip()
    tahun = _parse_tahun(request.POST.get('tahun', ''))
    if not tahap or tahun is None:
        messages.warning(request, 'Tahap dan tahun wajib diisi dengan benar')
        return redirect('desa_perhitungan_saw')
    if not Warga.objects.filter(desa=desa, is_deleted=False).exists():
        messages.warning(request, 'Data warga masih kosong')
        return redirect('desa_perhitungan_saw')
    if not Kriteria.objects.filter(desa=desa, aktif=True).exists():
        messages.warning(request, 'Belum ada kriteria yang diatur untuk desa ini')
        return redirect('desa_perhitungan_saw')
    if not jumlah_kpm:
        messages.warning(request, 'Jumlah KPM wajib diisi')
        return redirect('desa_perhitungan_saw')
    try:
        jumlah_kpm = int(jumlah_kpm)
        KuotaKPM.objects.update_or_create(desa=desa, tahap=tahap, tahun=tahun, defaults={'jumlah': jumlah_kpm})
    except ValueError:
        messages.warning(request, 'Jumlah KPM tidak valid') 
        return redirect('desa_perhitungan_saw')
    hitung_saw(desa=desa, tahap=tahap, tahun=tahun)

    messages.success(
        request,
        f'Proses SAW Tahap {tahap} Tahun {tahun} berhasil dijalankan.'
    )
    return redirect(
        f"{reverse('desa_perhitungan_saw')}?tahap={tahap}&tahun={tahun}"
    )


@login_required
@role_required('desa')
def riwayat_saw_list(request):
    desa = request.user.desa

    semua_periode = (
        HasilSAW.objects
        .filter(normalisasi__desa=desa)
        .values(
            'normalisasi__tahap',
            'normalisasi__tahun'
        )
        .distinct()
        .order_by(
            '-normalisasi__tahun',
            'normalisasi__tahap'
        )
    )
    periode_list = []
    for p in semua_periode:

        tahap = p['normalisasi__tahap']
        tahun = p['normalisasi__tahun']

        kuota = KuotaKPM.objects.filter(
            desa=desa,
            tahap=tahap,
            tahun=tahun
        ).first()

        ada_aktif = HasilSAW.objects.filter(
            normalisasi__desa=desa,
            normalisasi__tahap=tahap,
            normalisasi__tahun=tahun,
            normalisasi__is_active=True
        ).exists()

        periode_list.append({
            'tahap': tahap,
            'tahun': tahun,
            'jumlah_kpm': kuota.jumlah if kuota else '-',
            'is_active': ada_aktif,
        })

    return render(request, 'desa/riwayat_saw.html', {'periode_list': periode_list})


@login_required
@role_required('desa')
def hasil_saw_list(request):
    desa = request.user.desa

    tahap = request.GET.get('tahap', '').strip()
    tahun = _parse_tahun(request.GET.get('tahun', ''))

    hasil = HasilSAW.objects.none()
    normalisasi = NormalisasiSAW.objects.none()
    kriteria_list = Kriteria.objects.filter(desa=desa, aktif=True).order_by('id')
    kuota_obj = None
    sudah_diproses = False
    matriks_keputusan = []

    if tahap and tahun is not None:
        # Tabel HASIL / RANKING — tetap urut berdasarkan ranking (dari yang terbaik)
        hasil = (
            HasilSAW.objects
            .filter(
                normalisasi__desa=desa,
                normalisasi__tahap=tahap,
                normalisasi__tahun=tahun,
                normalisasi__is_active=True,
            )
            .select_related(
                'normalisasi',
                'normalisasi__warga'
            )
            .order_by('ranking')
        )

        # Tabel NORMALISASI — urut berdasarkan urutan data warga (id), bukan nama/ranking
        normalisasi = (
            NormalisasiSAW.objects
            .filter(desa=desa, tahap=tahap, tahun=tahun, is_active=True)
            .select_related('warga')
            .prefetch_related('detail__kriteria')
            .order_by('warga__id')
        )

        # Tabel MATRIKS KEPUTUSAN — juga urut berdasarkan urutan data warga (id)
        warga_list_urut = (
            Warga.objects
            .filter(
                desa=desa,
                is_deleted=False,
                normalisasisaw__tahap=tahap,
                normalisasisaw__tahun=tahun,
                normalisasisaw__is_active=True,
            )
            .order_by('id')
            .distinct()
        )

        warga_ids = [w.id for w in warga_list_urut]

        penilaian_qs = penilaianwarga.objects.filter(
            warga_id__in=warga_ids, kriteria_id__in=[k.id for k in kriteria_list]
        ).select_related('subkriteria', 'kriteria')

        nilai_mentah_map = {}
        for p in penilaian_qs:
            nilai_mentah_map.setdefault(p.warga_id, {})[p.kriteria_id] = p.subkriteria.nilai

        # Bangun matriks keputusan sesuai urutan warga (id), TERPISAH dari objek hasil
        for w in warga_list_urut:
            matriks_keputusan.append({
                'warga': w,
                'nilai_per_kriteria': [
                    nilai_mentah_map.get(w.id, {}).get(k.id, '—')
                    for k in kriteria_list
                ],
            })

        kuota_obj = KuotaKPM.objects.filter(desa=desa, tahap=tahap, tahun=tahun).first()
        sudah_diproses = hasil.exists()

    return render(request, 'desa/hasil_saw.html', {
        'hasil': hasil,
        'normalisasi': normalisasi,
        'matriks_keputusan': matriks_keputusan,
        'kriteria_list': kriteria_list,
        'sudah_diproses': sudah_diproses,
        'kuota': kuota_obj.jumlah if kuota_obj else '',
        'kuota_form': KuotaKPMForm(instance=kuota_obj),
        'tahap': tahap,
        'tahun': tahun if tahun else '',
    })


@login_required
@role_required('desa')
def set_kuota_kpm(request):
    if request.method == 'POST':
        desa = request.user.desa
        tahap = request.POST.get('tahap', '').strip()
        tahun = _parse_tahun(request.POST.get('tahun', ''))

        if not tahap or tahun is None:
            messages.warning(request, 'Tahap dan tahun wajib diisi dengan benar')
            return redirect('desa_riwayat_saw')

        kuota, _ = KuotaKPM.objects.get_or_create(desa=desa, tahap=tahap, tahun=tahun)
        form = KuotaKPMForm(request.POST, instance=kuota)
        if form.is_valid():
            form.save()
            messages.success(request, 'Kuota KPM berhasil disimpan')
        else:
            messages.warning(request, 'Data kuota tidak valid')

    return redirect(f"{reverse('desa_perhitungan_saw')}?tahap={tahap}&tahun={tahun}")




@login_required
@role_required('desa')
def hapus_saw(request):
    desa = request.user.desa

    tahap = request.POST.get('tahap', '').strip()
    tahun = _parse_tahun(request.POST.get('tahun', ''))

    if not tahap or tahun is None:
        messages.warning(request, 'Tahap dan tahun tidak valid')
        return redirect('desa_riwayat_saw')

    # Soft delete Normalisasi
    NormalisasiSAW.objects.filter(
        desa=desa,
        tahap=tahap,
        tahun=tahun,
        is_active=True
    ).update(
        is_active=False,
        deleted_at=timezone.now()
    )

    # Hapus hasil SAW
    HasilSAW.objects.filter(
        normalisasi__desa=desa,
        normalisasi__tahap=tahap,
        normalisasi__tahun=tahun
    ).delete()

    # Hapus kuota (opsional)
    KuotaKPM.objects.filter(
        desa=desa,
        tahap=tahap,
        tahun=tahun
    ).delete()

    messages.success(
        request,
        f'Data SAW Tahap {tahap} Tahun {tahun} berhasil dihapus.'
    )

    return redirect('desa_riwayat_saw')


@login_required
@role_required('desa')
def pengajuan_blt(request):
    desa = request.user.desa
    form = PengajuanBLTForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        pengajuan = form.save(commit=False)
        pengajuan.desa = desa
        pengajuan.save()
        messages.success(request, 'Pengajuan BLT berhasil dikirim')
        return redirect('desa_pengajuan_list')

    return render(request, 'desa/pengajuan_form.html', {'form': form})


@login_required
@role_required('desa')
def pengajuan_blt_list(request):
    desa = request.user.desa
    pengajuan = PengajuanBLT.objects.filter(desa=desa).order_by('-created_at')
    return render(request, 'desa/pengajuan_blt_list.html', {'pengajuan': pengajuan})


@login_required
@role_required('desa')
def pengajuan_blt_hapus(request, id):
    desa = request.user.desa
    pengajuan = get_object_or_404(PengajuanBLT, id=id, desa=desa)
    pengajuan.delete()
    messages.success(request, 'Data pengajuan BLT dihapus')
    return redirect('desa_pengajuan_list')


@login_required
@role_required('desa')
def pengajuan_blt_edit(request, id):
    desa = request.user.desa
    pengajuan = get_object_or_404(PengajuanBLT, id=id, desa=desa)
    form = PengajuanBLTForm(request.POST or None, request.FILES or None, instance=pengajuan)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Pengajuan BLT berhasil diperbarui')
        return redirect('desa_pengajuan_list')

    return render(request, 'desa/pengajuan_form.html', {'form': form})

@login_required
@role_required('desa')
def pengajuan_blt_detail(request, id):
    desa = request.user.desa
    pengajuan = get_object_or_404(PengajuanBLT, id=id, desa=desa)
    return render(request, 'desa/pengajuan_detail.html', {'pengajuan': pengajuan})



@login_required
@role_required('desa')
def export_hasil_saw_excel(request):
    desa = request.user.desa

    tahap = request.GET.get('tahap', '').strip()
    tahun = _parse_tahun(request.GET.get('tahun', ''))

    if not tahap or tahun is None:
        messages.warning(request, 'Tahap dan tahun wajib dipilih sebelum export')
        return redirect('desa_perhitungan_saw')

    # =========================
    # HASIL SAW
    # =========================
    hasil = (
        HasilSAW.objects
        .filter(
            normalisasi__desa=desa,
            normalisasi__tahap=tahap,
            normalisasi__tahun=tahun,
            normalisasi__is_active=True
        )
        .select_related(
            'normalisasi',
            'normalisasi__warga'
        )
        .prefetch_related(
            'normalisasi__detail__kriteria'
        )
        .order_by('ranking')
    )

    if not hasil.exists():
        messages.warning(
            request,
            f'Belum ada hasil SAW Tahap {tahap} Tahun {tahun}'
        )
        return redirect('desa_perhitungan_saw')

    # =========================
    # NORMALISASI
    # =========================
    normalisasi = (
        NormalisasiSAW.objects
        .filter(
            desa=desa,
            tahap=tahap,
            tahun=tahun,
            is_active=True
        )
        .select_related('warga')
        .prefetch_related('detail__kriteria')
    )

    # =========================
    # KRITERIA AKTIF
    # =========================
    kriteria_list = list(
        Kriteria.objects.filter(
            desa=desa,
            aktif=True
        ).order_by('urutan')
    )

    wb = Workbook()

    # =====================================================
    # SHEET 1 : HASIL SAW
    # =====================================================

    ws = wb.active
    ws.title = "Hasil SAW"

    headers = [
        "Ranking",
        "NIK",
        "Nama Warga",
        "Alamat"
    ]

    headers += [f"N {k.nama}" for k in kriteria_list]

    headers += [
        "Nilai Preferensi",
        "Desa",
        "Tanggal Proses"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for h in hasil:

        n = h.normalisasi

        detail_map = {
            d.kriteria_id: d.nilai_normalisasi
            for d in n.detail.all()
        }

        nama = (
            n.nama_warga_arsip
            or (n.warga.nama if n.warga else "Warga Terhapus")
        )

        nik = n.warga.nik if n.warga else "-"

        alamat = (
            n.warga.alamat
            if n.warga
            else "(Data warga sudah dihapus)"
        )

        row = [
            h.ranking,
            nik,
            nama,
            alamat,
        ]

        for k in kriteria_list:
            row.append(round(detail_map.get(k.id, 0), 4))

        row.extend([
            round(h.nilai_preferensi, 4),
            desa.nama_desa,
            h.tanggal_proses.strftime("%d-%m-%Y %H:%M")
            if h.tanggal_proses else "-"
        ])

        ws.append(row)

    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = length + 3

    # =====================================================
    # SHEET 2 : NORMALISASI
    # =====================================================

    ws2 = wb.create_sheet(title="Normalisasi SAW")

    headers = [
        "NIK",
        "Nama Warga"
    ]

    headers += [k.nama for k in kriteria_list]

    headers += [
        "Tanggal Proses"
    ]

    ws2.append(headers)

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for n in normalisasi:

        detail_map = {
            d.kriteria_id: d.nilai_normalisasi
            for d in n.detail.all()
        }

        nama = (
            n.nama_warga_arsip
            or (n.warga.nama if n.warga else "Warga Terhapus")
        )

        nik = n.warga.nik if n.warga else "-"

        row = [
            nik,
            nama
        ]

        for k in kriteria_list:
            row.append(round(detail_map.get(k.id, 0), 4))

        row.append(
            n.created_at.strftime("%d-%m-%Y %H:%M")
            if n.created_at else "-"
        )

        ws2.append(row)

    for col in ws2.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws2.column_dimensions[
            get_column_letter(col[0].column)
        ].width = length + 3

    # =====================================================
    # DOWNLOAD
    # =====================================================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = (
        f"Hasil_SAW_BLT_{desa.nama_desa}_Tahap_{tahap}_{tahun}.xlsx"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    wb.save(response)

    return response